import csv
import json
import os
import tempfile
import unittest

from modules.session_manager import SessionManager


class TestSessionManager(unittest.TestCase):
    def setUp(self):
        self._tmpdir = tempfile.mkdtemp()
        self.mgr = SessionManager(sessions_dir=self._tmpdir)

    def tearDown(self):
        import shutil
        self.mgr.end_session()
        shutil.rmtree(self._tmpdir, ignore_errors=True)

    def test_start_session_creates_file(self):
        meta = self.mgr.start_session()
        self.assertTrue(os.path.exists(os.path.join(self._tmpdir, meta.filename)))

    def test_start_session_writes_index(self):
        self.mgr.start_session()
        index_path = os.path.join(self._tmpdir, "index.json")
        self.assertTrue(os.path.exists(index_path))
        with open(index_path, encoding="utf-8") as f:
            data = json.load(f)
        self.assertEqual(len(data), 1)
        self.assertEqual(data[0]["fish_count"], 0)

    def test_append_catch_increments_count(self):
        self.mgr.start_session()
        self.mgr.append_catch("Salmon", "320")
        self.mgr.append_catch("Bass", "150")
        self.mgr.end_session()
        sessions = self.mgr.load_sessions()
        self.assertEqual(sessions[0].fish_count, 2)

    def test_append_catch_writes_csv_row(self):
        meta = self.mgr.start_session()
        self.mgr.append_catch("Salmon", "320")
        self.mgr.end_session()
        path = os.path.join(self._tmpdir, meta.filename)
        with open(path, newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f))
        self.assertEqual(rows[0], ["timestamp", "fish_name", "weight_g"])
        self.assertEqual(rows[1][1], "Salmon")
        self.assertEqual(rows[1][2], "320")

    def test_delete_session_removes_file_and_index_entry(self):
        meta = self.mgr.start_session()
        self.mgr.end_session()
        self.mgr.delete_session(meta.id)
        self.assertFalse(os.path.exists(os.path.join(self._tmpdir, meta.filename)))
        self.assertEqual(len(self.mgr.load_sessions()), 0)

    def test_load_sessions_rebuilds_from_files_if_index_missing(self):
        meta = self.mgr.start_session()
        self.mgr.append_catch("Trout", "200")
        self.mgr.end_session()
        index_path = os.path.join(self._tmpdir, "index.json")
        os.remove(index_path)
        sessions = self.mgr.load_sessions()
        self.assertEqual(len(sessions), 1)
        self.assertEqual(sessions[0].fish_count, 1)

    def test_export_csv(self):
        self.mgr.start_session()
        self.mgr.append_catch("Salmon", "320")
        meta = self.mgr.load_sessions()[0]
        self.mgr.end_session()
        out = os.path.join(self._tmpdir, "out.csv")
        self.mgr.export_session(meta.id, out, "csv")
        with open(out, newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f))
        self.assertEqual(rows[0], ["timestamp", "fish_name", "weight_g"])
        self.assertEqual(rows[1][1], "Salmon")

    def test_export_json(self):
        self.mgr.start_session()
        self.mgr.append_catch("Bass", "150")
        meta = self.mgr.load_sessions()[0]
        self.mgr.end_session()
        out = os.path.join(self._tmpdir, "out.json")
        self.mgr.export_session(meta.id, out, "json")
        with open(out, encoding="utf-8") as f:
            data = json.load(f)
        self.assertIsInstance(data, list)
        self.assertEqual(data[0]["fish_name"], "Bass")

    def test_export_xlsx(self):
        self.mgr.start_session()
        self.mgr.append_catch("Perch", "80")
        meta = self.mgr.load_sessions()[0]
        self.mgr.end_session()
        out = os.path.join(self._tmpdir, "out.xlsx")
        self.mgr.export_session(meta.id, out, "xlsx")
        import openpyxl
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, 4)]
        self.assertEqual(headers, ["timestamp", "fish_name", "weight_g"])
        self.assertEqual(ws.cell(2, 2).value, "Perch")


if __name__ == "__main__":
    unittest.main()
